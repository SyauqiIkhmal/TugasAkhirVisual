"""
Reporting Widget for Photo Studio Management System
Provides comprehensive reporting with PDF and Excel export capabilities
"""

import sys
import os
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                            QTableWidget, QTableWidgetItem, QHeaderView, QLabel,
                            QLineEdit, QDialog, QFormLayout, QDialogButtonBox,
                            QMessageBox, QFrame, QSplitter, QGroupBox, QComboBox,
                            QDateEdit, QTextEdit, QTabWidget, QProgressBar,
                            QFileDialog, QSpinBox)
from PyQt5.QtCore import Qt, pyqtSignal, QDate, QThread
from PyQt5.QtGui import QIcon, QPalette, QColor, QFont

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportGeneratorThread(QThread):
    """Thread for generating reports without blocking UI"""
    
    progress_updated = pyqtSignal(int)
    report_completed = pyqtSignal(str)
    error_occurred = pyqtSignal(str)
    
    def __init__(self, db_manager, report_type, start_date, end_date, output_path, format_type):
        super().__init__()
        self.db_manager = db_manager
        self.report_type = report_type
        self.start_date = start_date
        self.end_date = end_date
        self.output_path = output_path
        self.format_type = format_type
    
    def run(self):
        try:
            if self.format_type == "PDF":
                self.generate_pdf_report()
            elif self.format_type == "Excel":
                self.generate_excel_report()
        except Exception as e:
            self.error_occurred.emit(str(e))
    
    def generate_pdf_report(self):
        """Generate PDF report"""
        if not REPORTLAB_AVAILABLE:
            self.error_occurred.emit("ReportLab tidak tersedia. Install dengan: pip install reportlab")
            return
        
        self.progress_updated.emit(10)
        
        # Use landscape orientation for better table fit
        from reportlab.lib.pagesizes import landscape
        doc = SimpleDocTemplate(self.output_path, pagesize=landscape(A4), 
                              leftMargin=40, rightMargin=40, 
                              topMargin=60, bottomMargin=60)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        if self.report_type == "monthly":
            title = f"Laporan Bulanan - {self.start_date.strftime('%B %Y')}"
        else:
            title = f"Laporan Periode {self.start_date.strftime('%d/%m/%Y')} - {self.end_date.strftime('%d/%m/%Y')}"
        
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 12))
        
        self.progress_updated.emit(30)
        
        # Get data based on report type
        if self.report_type == "monthly":
            data = self.db_manager.get_monthly_report(
                self.start_date.year, self.start_date.month
            )
        else:
            data = self.get_period_report_data()
        
        self.progress_updated.emit(50)
        
        # Create simplified table with only essential columns
        table_data = [[
            'Tanggal/Waktu', 'Klien & Fotografer', 'Studio & Lokasi', 
            'Paket', 'Status'
        ]]
        
        for item in data:
            tanggal_waktu = item.get('tanggal_waktu', '')
            if isinstance(tanggal_waktu, str):
                try:
                    dt = datetime.fromisoformat(tanggal_waktu.replace('Z', '+00:00'))
                    formatted_dt = dt.strftime('%d/%m/%Y\n%H:%M')
                except:
                    formatted_dt = tanggal_waktu
            else:
                formatted_dt = str(tanggal_waktu)
            
            # Combine client and photographer
            client_name = item.get('nama_klien', '')
            photographer_name = item.get('nama_fotografer', '')
            if len(client_name) > 15:
                client_name = client_name.split()[0]
            if len(photographer_name) > 15:
                photographer_name = photographer_name.split()[0]
            client_photographer = f"{client_name}\n({photographer_name})"
            
            # Combine studio and location
            studio_name = item.get('nama_studio', '')
            location = item.get('lokasi', '')
            if len(studio_name) > 15:
                studio_name = studio_name.replace('Studio ', '')
            location_abbrev = location.replace('Jakarta ', 'JKT ').replace('Pusat', 'Pst').replace('Selatan', 'Sel').replace('Utara', 'Utr').replace('Barat', 'Brt').replace('Timur', 'Tmr')
            studio_location = f"{studio_name}\n{location_abbrev}"
            
            # Package type
            package = item.get('jenis_paket', '')
            
            table_data.append([
                formatted_dt,
                client_photographer,
                studio_location,
                package,
                item.get('status', '')
            ])
        
        self.progress_updated.emit(70)
        
        # Set column widths for landscape layout
        page_width = landscape(A4)[0] - 80  # Account for margins
        col_widths = [
            page_width * 0.18,  # Date/Time
            page_width * 0.28,  # Client & Photographer
            page_width * 0.28,  # Studio & Location
            page_width * 0.16,  # Package
            page_width * 0.10   # Status
        ]
        
        # Create and style table with proper column widths
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Status column centered
            
            # Layout and spacing
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('WORDWRAP', (0, 0), (-1, -1), True),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            
            # Minimal padding for compact design
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            
            # Row height optimization
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 12))
        
        # Add notes section if there are any important notes
        notes_data = [item for item in data if item.get('catatan', '').strip()]
        if notes_data:
            notes_style = ParagraphStyle(
                'Notes',
                parent=styles['Normal'],
                fontSize=8,
                spaceAfter=6
            )
            story.append(Paragraph("<b>Catatan Penting:</b>", notes_style))
            for i, item in enumerate(notes_data[:10], 1):  # Show max 10 notes
                client_name = item.get('nama_klien', '')
                date_str = ''
                try:
                    dt = datetime.fromisoformat(item.get('tanggal_waktu', '').replace('Z', '+00:00'))
                    date_str = dt.strftime('%d/%m/%Y')
                except:
                    pass
                note_text = f"{i}. {client_name} ({date_str}): {item.get('catatan', '')[:100]}"
                story.append(Paragraph(note_text, notes_style))
        
        self.progress_updated.emit(90)
        
        # Add summary
        story.append(Spacer(1, 12))
        summary_style = styles['Normal']
        
        total_sessions = len(data)
        booked_count = len([x for x in data if x.get('status') == 'Booked'])
        completed_count = len([x for x in data if x.get('status') == 'Selesai'])
        cancelled_count = len([x for x in data if x.get('status') == 'Batal'])
        
        summary_text = f"""
        <b>Ringkasan:</b><br/>
        Total Sesi: {total_sessions}<br/>
        Terjadwal: {booked_count}<br/>
        Selesai: {completed_count}<br/>
        Dibatalkan: {cancelled_count}<br/>
        <br/>
        Laporan dibuat pada: {datetime.now().strftime('%d/%m/%Y %H:%M')}
        """
        
        story.append(Paragraph(summary_text, summary_style))
        
        # Build PDF
        doc.build(story)
        
        self.progress_updated.emit(100)
        self.report_completed.emit(self.output_path)
    
    def generate_excel_report(self):
        """Generate Excel report"""
        if not OPENPYXL_AVAILABLE:
            self.error_occurred.emit("OpenPyXL tidak tersedia. Install dengan: pip install openpyxl")
            return
        
        self.progress_updated.emit(10)
        
        # Get data
        if self.report_type == "monthly":
            data = self.db_manager.get_monthly_report(
                self.start_date.year, self.start_date.month
            )
        else:
            data = self.get_period_report_data()
        
        self.progress_updated.emit(30)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if self.report_type == "monthly":
            ws.title = f"Laporan {self.start_date.strftime('%B %Y')}"
        else:
            ws.title = "Laporan Periode"
        
        # Headers
        headers = [
            'Tanggal/Waktu', 'Klien', 'Fotografer', 'Studio', 
            'Paket', 'Status', 'Catatan'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        self.progress_updated.emit(50)
        
        # Data rows
        for row, item in enumerate(data, 2):
            tanggal_waktu = item.get('tanggal_waktu', '')
            if isinstance(tanggal_waktu, str):
                try:
                    dt = datetime.fromisoformat(tanggal_waktu.replace('Z', '+00:00'))
                    formatted_dt = dt.strftime('%d/%m/%Y %H:%M')
                except:
                    formatted_dt = tanggal_waktu
            else:
                formatted_dt = str(tanggal_waktu)
            
            ws.cell(row=row, column=1, value=formatted_dt)
            ws.cell(row=row, column=2, value=item.get('nama_klien', ''))
            ws.cell(row=row, column=3, value=item.get('nama_fotografer', ''))
            ws.cell(row=row, column=4, value=f"{item.get('nama_studio', '')} - {item.get('lokasi', '')}")
            ws.cell(row=row, column=5, value=item.get('jenis_paket', ''))
            
            # Status with color
            status_cell = ws.cell(row=row, column=6, value=item.get('status', ''))
            status = item.get('status', '')
            if status == 'Booked':
                status_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            elif status == 'Selesai':
                status_cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
            elif status == 'Batal':
                status_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            ws.cell(row=row, column=7, value=item.get('catatan', ''))
        
        self.progress_updated.emit(80)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Ringkasan")
        total_sessions = len(data)
        booked_count = len([x for x in data if x.get('status') == 'Booked'])
        completed_count = len([x for x in data if x.get('status') == 'Selesai'])
        cancelled_count = len([x for x in data if x.get('status') == 'Batal'])
        
        summary_data = [
            ['Ringkasan Laporan', ''],
            ['Total Sesi', total_sessions],
            ['Terjadwal', booked_count],
            ['Selesai', completed_count],
            ['Dibatalkan', cancelled_count],
            ['', ''],
            ['Dibuat pada', datetime.now().strftime('%d/%m/%Y %H:%M')]
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row, column=2, value=value)
        
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 15
        
        self.progress_updated.emit(95)
        
        # Save workbook
        wb.save(self.output_path)
        
        self.progress_updated.emit(100)
        self.report_completed.emit(self.output_path)
    
    def get_period_report_data(self):
        """Get report data for custom period"""
        # This would need to be implemented in database_manager
        # For now, we'll get all data and filter
        all_data = self.db_manager.get_all_jadwal_with_details()
        filtered_data = []
        
        for item in all_data:
            tanggal_waktu_str = item.get('tanggal_waktu', '')
            if tanggal_waktu_str:
                try:
                    if isinstance(tanggal_waktu_str, str):
                        dt = datetime.fromisoformat(tanggal_waktu_str.replace('Z', '+00:00'))
                    else:
                        dt = tanggal_waktu_str
                    
                    if self.start_date <= dt.date() <= self.end_date:
                        filtered_data.append(item)
                except:
                    pass
        
        return filtered_data


class LaporanWidget(QWidget):
    """Main reporting widget"""
    
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setup_ui()
    
    def setup_ui(self):
        """Setup widget user interface"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)
        
        # Header section
        header_layout = QHBoxLayout()
        
        title_label = QLabel("Laporan & Ekspor Data")
        title_label.setStyleSheet("""
            font-size: 28px;
            font-weight: bold;
            color: #FFFFFF;
        """)
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        
        layout.addLayout(header_layout)
        
        # Main content area
        main_frame = QFrame()
        main_frame.setStyleSheet("""
            QFrame {
                background-color: #404040;
                border-radius: 10px;
                padding: 20px;
            }
        """)
        
        main_layout = QVBoxLayout(main_frame)
        main_layout.setSpacing(20)
        
        # Report generation section
        self.setup_report_section(main_layout)
        
        # Statistics section
        self.setup_statistics_section(main_layout)
        
        layout.addWidget(main_frame)
    
    def setup_report_section(self, parent_layout):
        """Setup report generation section"""
        report_group = QGroupBox("Generate Laporan")
        report_group.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #FFFFFF;
                border: 2px solid #505050;
                border-radius: 8px;
                margin: 5px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
            }
        """)
        
        report_layout = QVBoxLayout(report_group)
        report_layout.setSpacing(15)
        
        # Report type selection
        type_layout = QHBoxLayout()
        type_label = QLabel("Jenis Laporan:")
        type_label.setStyleSheet("color: #FFFFFF; font-weight: bold; font-size: 12px;")
        
        self.report_type_combo = QComboBox()
        self.report_type_combo.addItems(["Laporan Bulanan", "Laporan Periode Kustom"])
        self.report_type_combo.setStyleSheet("""
            QComboBox {
                background-color: #505050;
                color: #FFFFFF;
                border: 2px solid #606060;
                border-radius: 6px;
                padding: 8px;
                font-size: 12px;
                min-width: 200px;
            }
            QComboBox:focus {
                border-color: #4A90E2;
            }
        """)
        self.report_type_combo.currentTextChanged.connect(self.on_report_type_changed)
        
        type_layout.addWidget(type_label)
        type_layout.addWidget(self.report_type_combo)
        type_layout.addStretch()
        report_layout.addLayout(type_layout)
        
        # Date selection
        date_layout = QHBoxLayout()
        
        # Monthly selection
        self.monthly_frame = QFrame()
        monthly_layout = QHBoxLayout(self.monthly_frame)
        monthly_layout.setContentsMargins(0, 0, 0, 0)
        
        month_label = QLabel("Bulan:")
        month_label.setStyleSheet("color: #FFFFFF; font-weight: bold;")
        
        self.month_combo = QComboBox()
        months = [
            "Januari", "Februari", "Maret", "April", "Mei", "Juni",
            "Juli", "Agustus", "September", "Oktober", "November", "Desember"
        ]
        self.month_combo.addItems(months)
        self.month_combo.setCurrentIndex(datetime.now().month - 1)
        
        year_label = QLabel("Tahun:")
        year_label.setStyleSheet("color: #FFFFFF; font-weight: bold;")
        
        self.year_spin = QSpinBox()
        self.year_spin.setRange(2020, 2030)
        self.year_spin.setValue(datetime.now().year)
        
        for widget in [self.month_combo, self.year_spin]:
            widget.setStyleSheet("""
                QComboBox, QSpinBox {
                    background-color: #505050;
                    color: #FFFFFF;
                    border: 2px solid #606060;
                    border-radius: 6px;
                    padding: 8px;
                    font-size: 12px;
                }
                QComboBox:focus, QSpinBox:focus {
                    border-color: #4A90E2;
                }
            """)
        
        monthly_layout.addWidget(month_label)
        monthly_layout.addWidget(self.month_combo)
        monthly_layout.addWidget(year_label)
        monthly_layout.addWidget(self.year_spin)
        monthly_layout.addStretch()
        
        # Period selection
        self.period_frame = QFrame()
        period_layout = QHBoxLayout(self.period_frame)
        period_layout.setContentsMargins(0, 0, 0, 0)
        
        start_label = QLabel("Dari:")
        start_label.setStyleSheet("color: #FFFFFF; font-weight: bold;")
        
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setDate(QDate.currentDate().addDays(-30))
        self.start_date_edit.setCalendarPopup(True)
        
        end_label = QLabel("Sampai:")
        end_label.setStyleSheet("color: #FFFFFF; font-weight: bold;")
        
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setCalendarPopup(True)
        
        for widget in [self.start_date_edit, self.end_date_edit]:
            widget.setStyleSheet("""
                QDateEdit {
                    background-color: #505050;
                    color: #FFFFFF;
                    border: 2px solid #606060;
                    border-radius: 6px;
                    padding: 8px;
                    font-size: 12px;
                }
                QDateEdit:focus {
                    border-color: #4A90E2;
                }
            """)
        
        period_layout.addWidget(start_label)
        period_layout.addWidget(self.start_date_edit)
        period_layout.addWidget(end_label)
        period_layout.addWidget(self.end_date_edit)
        period_layout.addStretch()
        
        self.period_frame.hide()  # Initially hidden
        
        date_layout.addWidget(self.monthly_frame)
        date_layout.addWidget(self.period_frame)
        report_layout.addLayout(date_layout)
        
        # Export buttons
        export_layout = QHBoxLayout()
        
        self.export_pdf_btn = QPushButton("塘 Export ke PDF")
        self.export_pdf_btn.setStyleSheet("""
            QPushButton {
                background-color: #F44336;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #D32F2F;
            }
            QPushButton:disabled {
                background-color: #666666;
                color: #CCCCCC;
            }
        """)
        self.export_pdf_btn.clicked.connect(self.export_pdf)
        
        self.export_excel_btn = QPushButton("嶋 Export ke Excel")
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45A049;
            }
            QPushButton:disabled {
                background-color: #666666;
                color: #CCCCCC;
            }
        """)
        self.export_excel_btn.clicked.connect(self.export_excel)
        
        # Check library availability
        if not REPORTLAB_AVAILABLE:
            self.export_pdf_btn.setEnabled(False)
            self.export_pdf_btn.setToolTip("Install reportlab: pip install reportlab")
        
        if not OPENPYXL_AVAILABLE:
            self.export_excel_btn.setEnabled(False)
            self.export_excel_btn.setToolTip("Install openpyxl: pip install openpyxl")
        
        export_layout.addWidget(self.export_pdf_btn)
        export_layout.addWidget(self.export_excel_btn)
        export_layout.addStretch()
        
        report_layout.addLayout(export_layout)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #505050;
                border-radius: 8px;
                text-align: center;
                background-color: #2D2D2D;
                color: #FFFFFF;
            }
            QProgressBar::chunk {
                background-color: #4A90E2;
                border-radius: 6px;
            }
        """)
        self.progress_bar.hide()
        report_layout.addWidget(self.progress_bar)
        
        parent_layout.addWidget(report_group)
    
    def setup_statistics_section(self, parent_layout):
        """Setup statistics section"""
        stats_group = QGroupBox("Statistik Cepat")
        stats_group.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #FFFFFF;
                border: 2px solid #505050;
                border-radius: 8px;
                margin: 5px;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
            }
        """)
        
        stats_layout = QVBoxLayout(stats_group)
        
        # Quick stats display
        self.stats_frame = QFrame()
        self.stats_layout = QHBoxLayout(self.stats_frame)
        
        self.load_quick_stats()
        
        stats_layout.addWidget(self.stats_frame)
        
        # Refresh stats button
        refresh_stats_btn = QPushButton("売 Refresh Statistik")
        refresh_stats_btn.setStyleSheet("""
            QPushButton {
                background-color: #4A90E2;
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #357ABD;
            }
        """)
        refresh_stats_btn.clicked.connect(self.load_quick_stats)
        stats_layout.addWidget(refresh_stats_btn)
        
        parent_layout.addWidget(stats_group)
    
    def load_quick_stats(self):
        """Load quick statistics"""
        try:
            # Clear existing stats
            for i in reversed(range(self.stats_layout.count())):
                widget = self.stats_layout.itemAt(i).widget()
                if widget:
                    widget.setParent(None)
            
            # Check if db_manager exists
            if not hasattr(self, 'db_manager') or self.db_manager is None:
                error_label = QLabel("Database connection not available")
                error_label.setStyleSheet("color: #F44336; padding: 20px;")
                self.stats_layout.addWidget(error_label)
                return
            
            # Get dashboard stats
            stats = self.db_manager.get_dashboard_stats()
            
            # Create stat cards
            stat_items = [
                ("Total Klien", stats.get('total_klien', 0), "#4CAF50"),
                ("Total Fotografer", stats.get('total_fotografer', 0), "#2196F3"),
                ("Total Studio", stats.get('total_studio', 0), "#FF9800"),
                ("Sesi Terjadwal", stats.get('sesi_booked', 0), "#9C27B0"),
                ("Sesi Selesai", stats.get('sesi_selesai', 0), "#4CAF50"),
                ("Sesi Bulan Ini", stats.get('sesi_bulan_ini', 0), "#F44336")
            ]
            
            for label, value, color in stat_items:
                try:
                    stat_card = self.create_stat_card(label, value, color)
                    self.stats_layout.addWidget(stat_card)
                except Exception as card_error:
                    # FIX: Ubah format print error agar lebih aman dan tidak memicu error 'background'
                    print(f"Error creating stat card for {label}: {str(card_error)}")
            
        except Exception as e:
            print(f"Error in load_quick_stats: {e}")
            error_label = QLabel(f"Error loading stats: {str(e)}")
            error_label.setStyleSheet("color: #F44336; padding: 20px;")
            self.stats_layout.addWidget(error_label)
    
    def create_stat_card(self, label, value, color):
        """Create a statistics card widget"""
        card = QFrame()
        # FIX VISUAL: Menghilangkan border, menyesuaikan padding dan margin
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {color};
                border-radius: 8px;
                padding: 15px 10px; /* Tambah padding vertikal dan horizontal */
                margin: 0px 5px; /* Hapus margin vertikal, pertahankan margin horizontal antar kartu */
                border: none;
            }}
            QFrame:hover {{
                background-color: {color}E0; /* Efek hover sederhana */
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setAlignment(Qt.AlignCenter)
        
        value_label = QLabel(str(value))
        value_label.setStyleSheet("""
            font-size: 36px; /* Perbesar ukuran font */
            font-weight: 800; /* Lebih tebal */
            color: white;
            padding-bottom: 5px;
        """)
        value_label.setAlignment(Qt.AlignCenter)
        
        desc_label = QLabel(label)
        desc_label.setStyleSheet("""
            font-size: 14px; /* Perbesar ukuran font deskripsi */
            color: #E0E0E0; /* Warna putih sedikit redup */
            font-weight: 500;
        """)
        desc_label.setAlignment(Qt.AlignCenter)
        
        layout.addWidget(value_label)
        layout.addWidget(desc_label)
        
        return card
    
    def on_report_type_changed(self, report_type):
        """Handle report type change"""
        if report_type == "Laporan Bulanan":
            self.monthly_frame.show()
            self.period_frame.hide()
        else:
            self.monthly_frame.hide()
            self.period_frame.show()
    
    def export_pdf(self):
        """Export report to PDF"""
        self.export_report("PDF")
    
    def export_excel(self):
        """Export report to Excel"""
        self.export_report("Excel")
    
    def export_report(self, format_type):
        """Export report in specified format"""
        try:
            # Get report parameters
            report_type = "monthly" if self.report_type_combo.currentText() == "Laporan Bulanan" else "period"
            
            if report_type == "monthly":
                month = self.month_combo.currentIndex() + 1
                year = self.year_spin.value()
                start_date = datetime(year, month, 1).date()
                end_date = start_date  # Not used for monthly
                filename_date = f"{year}_{month:02d}"
            else:
                start_date = self.start_date_edit.date().toPyDate()
                end_date = self.end_date_edit.date().toPyDate()
                filename_date = f"{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"
            
            # Get output file path
            extension = "pdf" if format_type == "PDF" else "xlsx"
            default_filename = f"laporan_{filename_date}.{extension}"
            
            file_filter = f"{format_type} Files (*.{extension})"
            output_path, _ = QFileDialog.getSaveFileName(
                self, f"Save {format_type} Report", default_filename, file_filter
            )
            
            if not output_path:
                return
            
            # Show progress bar
            self.progress_bar.show()
            self.progress_bar.setValue(0)
            
            # Disable export buttons
            self.export_pdf_btn.setEnabled(False)
            self.export_excel_btn.setEnabled(False)
            
            # Start report generation thread
            self.report_thread = ReportGeneratorThread(
                self.db_manager, report_type, start_date, end_date, output_path, format_type
            )
            self.report_thread.progress_updated.connect(self.progress_bar.setValue)
            self.report_thread.report_completed.connect(self.on_report_completed)
            self.report_thread.error_occurred.connect(self.on_report_error)
            self.report_thread.start()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memulai export: {str(e)}")
    
    def on_report_completed(self, file_path):
        """Handle report completion"""
        self.progress_bar.hide()
        self.export_pdf_btn.setEnabled(REPORTLAB_AVAILABLE)
        self.export_excel_btn.setEnabled(OPENPYXL_AVAILABLE)
        
        QMessageBox.information(
            self, "Export Berhasil",
            f"Laporan berhasil disimpan ke:\n{file_path}"
        )
    
    def on_report_error(self, error_message):
        """Handle report error"""
        self.progress_bar.hide()
        self.export_pdf_btn.setEnabled(REPORTLAB_AVAILABLE)
        self.export_excel_btn.setEnabled(OPENPYXL_AVAILABLE)
        
        QMessageBox.critical(
            self, "Export Gagal",
            f"Gagal membuat laporan:\n{error_message}"
        )